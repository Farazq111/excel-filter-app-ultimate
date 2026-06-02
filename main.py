import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==========================================
# DICTIONARY FOR STATE MAPPING
# ==========================================
STATE_MAPING = {
    "RJ": "Rajasthan", "MH": "Maharashtra", "BR": "Bihar", "GJ": "Gujarat",
    "MP": "Madhya Pradesh", "KA": "Karnataka", "NE": "North East", "JK": "Jammu & Kashmir",
    "UP": "Uttar Pradesh", "DL": "Delhi", "HR": "Haryana", "PB": "Punjab",
    "TN": "Tamil Nadu", "AP": "Andhra Pradesh", "WB": "West Bengal", "OR": "Odisha",
    "AS": "Assam", "UE": "Uttar Pradesh East", "UW": "Uttar Pradesh West",
    "KL": "Kerala", "CG": "Chhattisgarh",
}

# ==========================================
# HELPER FUNCTION: OWNER LOGIC
# ==========================================
def determine_owner(sap_id):
    if pd.isna(sap_id):
        return None
    
    parts = str(sap_id).strip().split('-')
    last_part = parts[-1] if parts else ""
    
    if any(c.isalpha() for c in last_part):
        return "P1 colo"
    elif last_part.isdigit() and len(last_part) > 0:
        first_digit = last_part[0]
        if first_digit in ['6', '7', '8']:
            return "RP1"
        else:
            return "P1"
    else:
        return "Unknown"

# ==========================================
# MAIN PROCESSING FUNCTION
# ==========================================
def process_data(path_mfl_dir, path_ht_dir, path_dg_dir, output_dir, status_lbl, root_win):
    try:
        OUTPUT_FILE = os.path.join(output_dir, "Final-marge-data2.xlsx")
        
        # UI Feedback Update
        status_lbl.config(text="Processing started... Checking folders.", fg="#ffcc00")
        root_win.update()
        
        # ------------------------------------------
        # PHASE 2: PROCESSING SOC DATA (MFL FILES)
        # ------------------------------------------
        status_lbl.config(text="Processing SOC (MFL) Files...", fg="#ffcc00")
        root_win.update()
        
        COLS_SOC = ["SAP ID", "Circle", "Maintaince Point", "Alarm Type", "Alarm Dated", "Facility", "Owner"]
        all_soc_dfs = []
        search_mfl = os.path.join(path_mfl_dir, "MFL*.xlsx")
        
        for file in glob.glob(search_mfl):
            try:
                df = pd.read_excel(file, sheet_name="Worksheet")[COLS_SOC[:-1]]
                df = df[df["Alarm Type"].str.contains("SOC low|DC voltage", case=False, na=False)].copy()
                df.loc[df["Alarm Type"].str.contains("SOC low", case=False, na=False), "Alarm Type"] = "SOC LOW"
                df["Owner"] = df["SAP ID"].apply(determine_owner)
                all_soc_dfs.append(df)
            except Exception as e:
                print(f"Error processing SOC file {os.path.basename(file)}: {e}")

        # ------------------------------------------
        # PHASE 3: PROCESSING HT DATA
        # ------------------------------------------
        status_lbl.config(text="Processing HT Files...", fg="#ffcc00")
        root_win.update()
        
        COLS_HT = ["NUMBER", "ASSIGNMENT", "OPEN_TIME", "TITLE", "SAP ID", "Owner", "state", "site name"]
        all_ht_dfs = []
        search_ht = os.path.join(path_ht_dir, "Export-*.xlsx")
        
        for file in glob.glob(search_ht):
            try:
                df = pd.read_excel(file, sheet_name="Data")
                condition_high_temp = df["TITLE"].str.contains("High temp", case=False, na=False)
                condition_keywords = df["TITLE"].str.contains("collector|core|Ciena", case=False, na=False)
                condition_no_metro_core = ~df["TITLE"].str.contains("metro core", case=False, na=False)
                
                df = df[condition_high_temp & condition_keywords & condition_no_metro_core].copy()
                
                if not df.empty:
                    sap_id_pattern = r"([A-Za-z0-9]-[A-Za-z0-9]{2}-[A-Za-z0-9]{4}-[A-Za-z0-9]{3}-[0-9]{4})"
                    df["SAP ID"] = df["TITLE"].str.extract(sap_id_pattern)
                    df["SAP ID"] = df["SAP ID"].str.strip()
                    df["Owner"] = df["SAP ID"].apply(determine_owner)
                    df["state"] = df["SAP ID"].str.extract(r"^[A-Za-z0-9]-([A-Za-z0-9]{2})-")
                    
                    missing_state = df["state"].isna()
                    if missing_state.any():
                        df.loc[missing_state, "state"] = df.loc[missing_state, "TITLE"].str.extract(r"(?i)circle\s*:\s*([A-Za-z]{2})")
                    
                    fallback_state_pattern = r"(?i)(?://|/|\|\||\|)\s*([A-Za-z]{2})\s*(?://|/|\|\||\|)"
                    missing_state = df["state"].isna()
                    if missing_state.any():
                        df.loc[missing_state, "state"] = df.loc[missing_state, "TITLE"].str.extract(fallback_state_pattern)
                    
                    df["state"] = df["state"].str.upper().str.strip()
                    df["state"] = df["state"].map(STATE_MAPING).fillna(df["state"])

                    df["site name"] = df["TITLE"].str.extract(r"(?i)site(?:\s+name)?\s*:\s*/*([^/|]+)")
                    fallback_site_pattern = r"(?i)(?://|/|\|\||\|)\s*([^/|]+?)\s*(?://|/|\|\||\|)\s*(?:core|collector|Ciena)"
                    missing_site = df["site name"].isna()
                    if missing_site.any():
                        df.loc[missing_site, "site name"] = df.loc[missing_site, "TITLE"].str.extract(fallback_site_pattern)
                        
                    df["site name"] = df["site name"].str.strip().str.strip('/')
                    df["site name"] = df["site name"].str.strip()
                    
                    df = df[COLS_HT]
                    all_ht_dfs.append(df)
            except Exception as e:
                print(f"Error processing HT file {os.path.basename(file)}: {e}")

        # ------------------------------------------
        # PHASE 4: PROCESSING DG IN MANUAL DATA
        # ------------------------------------------
        status_lbl.config(text="Processing DG Manual Files...", fg="#ffcc00")
        root_win.update()
        
        COLS_DG = ['alarm_name', 'sapid', 'owner', 'state', 'alarm_raised_time', 'rj_site_owner', 'sitetype']
        all_dg_dfs = []
        
        for path in glob.glob(os.path.join(path_dg_dir, '**', '*.csv'), recursive=True):
            fname = os.path.basename(path).lower()
            try:
                df = pd.read_csv(path, encoding='latin-1')
                df.columns = df.columns.str.strip()
                
                available_cols = [c for c in df.columns if c.lower() in COLS_DG]
                df = df[available_cols].copy()
                df.columns = df.columns.str.lower()
                
                if 'sapid' in df.columns:
                    df["owner"] = df["sapid"].apply(determine_owner)
                    df["state"] = df["sapid"].str.extract(r"^[A-Za-z0-9]-([A-Za-z0-9]{2})-")
                    
                    missing_state = df["state"].isna()
                    if missing_state.any() and 'alarm_name' in df.columns:
                        fallback_state_pattern = r"(?i)(?://|/|\|\||\|)\s*([A-Za-z]{2})\s*(?://|/|\|\||\|)"
                        df.loc[missing_state, "state"] = df.loc[missing_state, "alarm_name"].str.extract(fallback_state_pattern)
                    
                    df["state"] = df["state"].str.upper().str.strip()
                    df["state"] = df["state"].map(STATE_MAPING).fillna(df["state"])
                else:
                    df["owner"] = None
                    df["state"] = None
                    
                if 'lfl' in fname:
                    df = df[[c for c in COLS_DG if c in df.columns]]
                    all_dg_dfs.append(df)
                elif 'sfl' in fname and 'sitetype' in df.columns:
                    df = df[df['sitetype'].astype(str).str.upper().str.strip() == 'ILA']
                    if not df.empty:
                        df = df[[c for c in COLS_DG if c in df.columns]]
                        all_dg_dfs.append(df)
            except Exception as e:
                print(f"Error processing DG file {fname}: {e}")

        # ------------------------------------------
        # PHASE 5: CONSOLIDATION & FORMATTING
        # ------------------------------------------
        if not all_soc_dfs and not all_ht_dfs and not all_dg_dfs:
            status_lbl.config(text="No valid data found.", fg="#ff3333")
            messagebox.showwarning("No Data", "Selected folders mein koi valid files nahi mili!")
            return

        status_lbl.config(text="Generating & Formatting Excel...", fg="#ffcc00")
        root_win.update()

        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
            if all_soc_dfs:
                pd.concat(all_soc_dfs, ignore_index=True).to_excel(writer, sheet_name="SOC", index=False)
            if all_ht_dfs:
                pd.concat(all_ht_dfs, ignore_index=True).to_excel(writer, sheet_name="HT", index=False)
            if all_dg_dfs:
                pd.concat(all_dg_dfs, ignore_index=True).to_excel(writer, sheet_name="DG in Manual", index=False)

            workbook = writer.book
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            thin_side = Side(border_style="thin", color="000000")
            cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = yellow_fill
                    cell.border = cell_border
                    
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = cell_border

        status_lbl.config(text="Successful !", fg="#00ff00")
        messagebox.showinfo("Success", f"File successfully generated ho gayi hai!\nPath: {OUTPUT_FILE}")
        
    except Exception as e:
        status_lbl.config(text="Error Occurred", fg="#ff3333")
        messagebox.showerror("Error", f"Kuch gadbad ho gayi:\n{str(e)}")

# ==========================================
# GUI SETUP (CLEAN & SLEEK DARK MODE)
# ==========================================
def start_gui():
    root = tk.Tk()
    root.title("ILA OADM Report - Advanced Data Merger")
    root.geometry("650x450")
    root.resizable(False, False)
    root.configure(bg="#212121")

    # Path Storage Variables
    mfl_path = tk.StringVar()
    ht_path = tk.StringVar()
    dg_path = tk.StringVar()
    out_path = tk.StringVar()

    # Browse Trigger Functions
    def browse_mfl(): mfl_path.set(filedialog.askdirectory(title="Select MFL Files Folder"))
    def browse_ht(): ht_path.set(filedialog.askdirectory(title="Select HT Files Folder"))
    def browse_dg(): dg_path.set(filedialog.askdirectory(title="Select DG Manual Folder"))
    def browse_out(): out_path.set(filedialog.askdirectory(title="Select Output Folder Save Location"))

    def run_process():
        if not mfl_path.get() or not ht_path.get() or not dg_path.get() or not out_path.get():
            messagebox.showwarning("Input Required", "Kripya karke saare folders select karein!")
            return
        process_data(mfl_path.get(), ht_path.get(), dg_path.get(), out_path.get(), status_label, root)

    # Styles & UI Components
    tk.Label(root, text="Upload SOC, DG & HT Dumps", font=("Arial", 20, "bold"), bg="#212121", fg="#ffffff").pack(pady=(30, 20))

    # Helper function to generate symmetric dark rows
    def create_input_row(label_text, text_var, browse_cmd):
        frame = tk.Frame(root, bg="#212121")
        frame.pack(fill="x", padx=30, pady=8)
        
        lbl = tk.Label(frame, text=label_text, width=18, anchor="w", bg="#212121", fg="#ffffff", font=("Arial", 10, "bold"))
        lbl.pack(side="left")
        
        entry = tk.Entry(frame, textvariable=text_var, width=40, bg="#333333", fg="#ffffff", insertbackground="white", bd=1, relief="solid")
        entry.pack(side="left", padx=10, ipady=3)
        
        btn = tk.Button(frame, text="Browse", command=browse_cmd, width=10, bg="#424242", fg="white", activebackground="#555555", activeforeground="white", bd=0, relief="flat", cursor="hand2")
        btn.pack(side="left")

    # Creating UI Input Sections
    create_input_row("MFL Folder:", mfl_path, browse_mfl)
    create_input_row("HT Folder:", ht_path, browse_ht)
    create_input_row("DG Manual Folder:", dg_path, browse_dg)
    create_input_row("Output Save Folder:", out_path, browse_out)

    # Master Execution Trigger Button (Custom Styled Blue)
    submit_btn = tk.Button(
        root,
        text="START MERGE PROCESS",
        font=("Arial", 12, "bold"),
        bg="#1f6aa5",
        fg="white",
        activebackground="#144871",
        activeforeground="white",
        bd=0,
        padx=40,
        pady=10,
        cursor="hand2",
        command=run_process
    )
    submit_btn.pack(pady=(25, 10))

    # Dynamic Live System Processing Status Label
    status_label = tk.Label(root, text="", font=("Arial", 11, "bold"), bg="#212121", fg="#ffffff")
    status_label.pack()

    root.mainloop()

if __name__ == "__main__":
    start_gui()
